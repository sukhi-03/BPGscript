import os
import requests
import pandas as pd
from urllib.parse import quote
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
import subprocess
from tqdm import tqdm
import fitz  # PyMuPDF
from openpyxl import Workbook, load_workbook
import os


# -------- Parse BPG string --------
def parse_bpg(bpg_str):
    parts = bpg_str.split("~")
    return {
        "BIN": None if parts[0] == "NULL" else parts[0],
        "PCN": None if parts[1] == "NULL" else parts[1],
        "GroupID": None if parts[2] == "NULL" else parts[2],
    }

# -------- Search Google for PDFs --------
def search_pdf_links(query, max_results=3):
    links = []
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36")
            page = context.new_page()

            try:
                page.goto(f"https://www.bing.com/search?q={quote(query)}", timeout=30000)
                page.wait_for_selector("li.b_algo a", timeout=10000)
            except PlaywrightTimeoutError:
                print(f"⚠️ Timeout during Bing search for: {query}")
                return []

            anchors = page.query_selector_all("li.b_algo a")
            for a in anchors:
                href = a.get_attribute("href")
                if href and ".pdf" in href.lower() and href.startswith("http"):
                    links.append(href)
                    if len(links) >= max_results:
                        break

            browser.close()
    except Exception as e:
        print(f"❌ Error in search_pdf_links(): {e}")
        return []

    return links

# -------- Download PDF --------
def download_pdf(url):
    try:
        os.makedirs("pdfs", exist_ok=True)
        from pathlib import Path
        filename = Path("pdfs") / Path(url).name.split("?")[0]

        if os.path.exists(filename):
            return filename
        r = requests.get(url, timeout=10)
        with open(filename, "wb") as f:
            f.write(r.content)
        return filename
    except Exception as e:
        print(f"Download failed for {url}: {e}")
        return None

# -------- Extract Text from PDF --------
def extract_text_from_pdf(path):
    text = ""
    try:
        if not os.path.exists(path):
            print(f"❌ PDF file not found: {path}")
            return ""

        doc = fitz.open(path)
        for page in doc:
            text += page.get_text()
        return text[:6000]  # truncate for LLM prompt size limit
    except Exception as e:
        print(f"Failed to extract text from {path}: {e}")
        return ""

# -------- Ask Ollama to Extract Info --------
def ask_ollama(text, expected_bin, expected_pcn, expected_group):
    prompt = f"""
You are a data extraction expert. From the following PDF content, extract any BIN, PCN, Group ID, and Plan Type info you find.

Then compare the extracted values to these expected inputs:

Expected BIN: {expected_bin or "N/A"}
Expected PCN: {expected_pcn or "N/A"}
Expected Group ID: {expected_group or "N/A"}

Only return rows that either match or closely resemble the expected values. If there's no match, write that in the Comments.

Return only a table with columns:

BIN | PCN | Group ID | Plan type | Comments

Text:
{text}
"""
    try:
        result = subprocess.run(
            ["ollama", "run", "llama3.1:8b-instruct-q2_K"],
            input=prompt.encode(),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=300
        )
        if result.returncode != 0:
            print(f"Ollama error: {result.stderr.decode()}")
            return ""
        return result.stdout.decode()
    except Exception as e:
        print(f"Ollama call failed: {e}")
        return ""

# -------- Parse LLM Output --------
def parse_llm_output(output):
    rows = []
    for line in output.splitlines():
        if "|" in line:
            parts = [p.strip() for p in line.split("|")]
            while len(parts) < 5:
                parts.append("")  # Fill missing fields
            parts.append("")  # PDF link
            rows.append(parts)
    return rows

# -------- Main --------
def main():
    df = pd.read_excel("BPG.xlsx", header=None)
    df.columns = ["BPG"]
    df = df[::-1].reset_index(drop=True)

    output_path = "BPG_output_CT.xlsx"

    # Load existing file or create a new workbook
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["BIN", "PCN", "Group ID", "Plan Type", "Comments", "PDF Link", "Status"])
        wb.save(output_path)

    for idx, row in tqdm(df.iterrows(), total=len(df), desc="Processing BPGs"):
        bpg_data = parse_bpg(row["BPG"])
        query = " ".join([
            f"BIN {bpg_data['BIN']}" if bpg_data['BIN'] else "",
            f"PCN {bpg_data['PCN']}" if bpg_data['PCN'] else "",
            f"Group ID {bpg_data['GroupID']}" if bpg_data['GroupID'] else "",
            "filetype:pdf"
        ]).strip()

        tqdm.write(f"[{idx+1}] Searching for: {query}")

        pdf_links = search_pdf_links(query)
        if not pdf_links:
            ws.append(["", "", bpg_data['GroupID'] or "", "", "No PDF found", "", "❌ No PDF"])
            wb.save(output_path)
            continue

        found_valid_pdf = False
        for link in pdf_links:
            pdf_path = download_pdf(link)
            if not pdf_path:
                continue

            text = extract_text_from_pdf(pdf_path)
            if not text:
                continue

            llm_output = ask_ollama(text, bpg_data['BIN'], bpg_data['PCN'], bpg_data['GroupID'])  # pass expected values
            rows = parse_llm_output(llm_output)

            for r in rows:
                r[5] = link  # Add PDF link
                if all([r[0], r[1], r[2], r[3]]):
                    status = "✅ Success"
                elif any([r[0], r[1], r[2], r[3]]):
                    status = "⚠️ Partial extraction"
                else:
                    status = "⚠️ No usable info"
                r.append(status)
                ws.append(r)
                wb.save(output_path)

            found_valid_pdf = True
            break  # Stop after first valid PDF

        if not found_valid_pdf:
            ws.append(["", "", bpg_data['GroupID'] or "", "", "No valid PDF content extracted", "", "⚠️ No usable info"])
            wb.save(output_path)

    print(f"✅ Done. Results saved to {output_path}")

if __name__ == "__main__":
    main()
