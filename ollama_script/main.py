import os
import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook, load_workbook
import os
from parse import parse_bpg, parse_llm_output
from ask_ollama import ask_ollama
from pdf import search_pdf_links, download_pdf, extract_text_from_pdf

# -------- Main --------
# def main():
#     df = pd.read_excel("./input/BPG.xlsx", header=None)
#     df.columns = ["BPG"]
#     df = df[::-1].reset_index(drop=True)

#     output_path = "./output/BPG_output_CT.xlsx"

#     # Load existing file or create a new workbook
#     if os.path.exists(output_path):
#         wb = load_workbook(output_path)
#         ws = wb.active
#     else:
#         wb = Workbook()
#         ws = wb.active
#         ws.append(["BIN", "PCN", "Group ID", "Plan Type", "Comments", "PDF Link", "Status"])
#         wb.save(output_path)

#     for idx, row in tqdm(df.iterrows(), total=len(df), desc="Processing BPGs"):
#         bpg_data = parse_bpg(row["BPG"])
#         query = " ".join([
#             f"BIN {bpg_data['BIN']}" if bpg_data['BIN'] else "",
#             f"PCN {bpg_data['PCN']}" if bpg_data['PCN'] else "",
#             f"Group ID {bpg_data['GroupID']}" if bpg_data['GroupID'] else "",
#             "filetype:pdf"
#         ]).strip()

#         tqdm.write(f"[{idx+1}] Searching for: {query}")

#         pdf_links = search_pdf_links(query)
#         if not pdf_links:
#             ws.append(["", "", bpg_data['GroupID'] or "", "", "No PDF found", "", "❌ No PDF"])
#             wb.save(output_path)
#             continue

#         found_valid_pdf = False
#         for link in pdf_links:
#             pdf_path = download_pdf(link)
#             if not pdf_path:
#                 continue

#             text = extract_text_from_pdf(pdf_path)
#             if not text:
#                 continue

#             llm_output = ask_ollama(text, bpg_data['BIN'], bpg_data['PCN'], bpg_data['GroupID'])  # pass expected values
#             rows = parse_llm_output(llm_output)

#             for r in rows:
#                 r[5] = link  # Add PDF link
#                 if all([r[0], r[1], r[2], r[3]]):
#                     status = "✅ Success"
#                 elif any([r[0], r[1], r[2], r[3]]):
#                     status = "⚠️ Partial extraction"
#                 else:
#                     status = "⚠️ No usable info"
#                 r.append(status)
#                 ws.append(r)
#                 wb.save(output_path)

#             found_valid_pdf = True
#             break  # Stop after first valid PDF

#         if not found_valid_pdf:
#             ws.append(["", "", bpg_data['GroupID'] or "", "", "No valid PDF content extracted", "", "⚠️ No usable info"])
#             wb.save(output_path)

#     print(f"✅ Done. Results saved to {output_path}")


#===================
# line-by-line
# ==================
def main():
    df = pd.read_excel("./input/BPG.xlsx", header=None)
    df.columns = ["BPG"]
    # df = df[::-1].reset_index(drop=True)
    df = df[::-1].reset_index(drop=True).iloc[:5]  # Only first 5


    output_path = "./output/BPG_output_CT.xlsx"

    # Load existing file or create a new workbook
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["BIN", "PCN", "GroupID", "Plan Type", "Comments", "PDF Link", "Status"])
        wb.save(output_path)

    for idx, row in tqdm(df.iterrows(), total=len(df), desc="Processing BPGs"):
        bpg_data = parse_bpg(row["BPG"])
        query = " ".join([
            f"BIN {bpg_data['BIN']}" if bpg_data['BIN'] else "",
            f"PCN {bpg_data['PCN']}" if bpg_data['PCN'] else "",
            f"GroupID {bpg_data['GroupID']}" if bpg_data['GroupID'] else "",
            "filetype:pdf"
        ]).strip()

        tqdm.write(f"\n[{idx + 1}] 🔍 Searching for: {query}")

        try:
            pdf_links = search_pdf_links(query)
        except Exception as e:
            tqdm.write(f"❌ Error during search: {e}")
            ws.append([bpg_data['BIN'], bpg_data['PCN'], bpg_data['GroupID'], "", "Search error", "", "❌ Search failed"])
            wb.save(output_path)
            continue

        if not pdf_links:
            ws.append([bpg_data['BIN'], bpg_data['PCN'], bpg_data['GroupID'], "", "No PDFs found", "", "❌ No PDF"])
            wb.save(output_path)
            continue

        processed = False
        for link in pdf_links:
            tqdm.write(f"➡️  Trying PDF: {link}")
            pdf_path = download_pdf(link)
            if not pdf_path:
                tqdm.write("⚠️  PDF download failed.")
                continue

            text = extract_text_from_pdf(pdf_path)
            if not text:
                tqdm.write("⚠️  PDF text extraction failed.")
                continue

            llm_output = ask_ollama(text, bpg_data['BIN'], bpg_data['PCN'], bpg_data['GroupID'])
            if not llm_output.strip():
                tqdm.write("⚠️  LLM output was empty.")
                continue

            rows = parse_llm_output(llm_output)
            if not rows:
                tqdm.write("⚠️  LLM output could not be parsed.")
                continue

            for r in rows:
                r[5] = link  # PDF link
                if all([r[0], r[1], r[2], r[3]]):
                    status = "✅ Success"
                elif any([r[0], r[1], r[2], r[3]]):
                    status = "⚠️ Partial extraction"
                else:
                    status = "⚠️ No usable info"
                r.append(status)
                ws.append(r)

            wb.save(output_path)
            processed = True
            break  # Exit after first valid PDF

        if not processed:
            ws.append([bpg_data['BIN'], bpg_data['PCN'], bpg_data['GroupID'], "", "No usable PDF found", "", "⚠️ No usable info"])
            wb.save(output_path)

    print(f"\n✅ All BPGs processed. Results saved to: {output_path}")

if __name__ == "__main__":
    main()
